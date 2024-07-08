import os.path

import openpyxl
import pandas as pd


class DataReader:
    def __init__(self, file_path, file_name):
        self.file_path = file_path
        self.file_name = file_name

    def is_file_path_exists(self):
        color = 'red'

        if not self.file_path:
            msg = "Путь к файлу не задан!"
            return False, msg, color

        if not os.path.exists(self.file_path):
            msg = f"Путь {self.file_path} не найден!"
            return False, msg, color

        if not self.file_name:
            msg = "Файл не задан!"
            return False, msg, color

        try:
            path = os.path.abspath(os.path.join(self.file_path, self.file_name))
        except Exception as e:
            msg = f'ОШИБКА {type(e)}: {e}'
            return False, msg, color

        if not os.path.exists(path):
            msg = f'Файл {self.file_name} по пути {self.file_path} не найден!'
            color = 'orange'
            return False, msg, color

        msg = ''
        color = 'blue'
        return True, msg, color

    def read_csv_file(self, column_name):

        is_file_path_exists, msg, color = self.is_file_path_exists()
        data = None

        if is_file_path_exists:
            path = os.path.abspath(os.path.join(self.file_path, self.file_name))

            try:
                data = pd.read_csv(path).sort_values(by=column_name)  #, ascending=False
                color = 'blue'
            except Exception as e:
                msg = f'ОШИБКА {type(e)}: {e}'
                color = 'red'

        return data, msg, color

    def add_to_file_csv(self, df):
        # df already validated

        is_file_path_exists, msg, color = self.is_file_path_exists()

        if color != 'red':
            path = os.path.abspath(os.path.join(self.file_path, self.file_name))

            if is_file_path_exists:
                df.to_csv(path, mode='a', index=False, header=None)
                msg = f'Запись добавлена в файл {self.file_name} !'
            else:
                if color == 'orange':
                    df.to_csv(path, mode='a', index=False)
                    msg = f'Файл {self.file_name} создан! Запись добавлена!'
                    color = 'blue'

        return msg, color

    def read_data_from_stock_file(self, sheet_name, cols):
        is_file_path_exists, msg, color = self.is_file_path_exists()
        # print(is_file_path_exists, msg, color)
        data = None

        if is_file_path_exists:
            path = os.path.abspath(os.path.join(self.file_path, self.file_name))

        # if self.path_to_file is not None and os.path.exists(f'{self.path_to_file}{self.file_name}'):
            try:
                # self.data = pd.read_excel(
                data = pd.read_excel(
                    # f'{self.path_to_file}{self.file_name}',
                    path,
                    sheet_name=sheet_name,
                    usecols=cols,
                    engine='openpyxl',
                )
                # self.color = 'blue'
                color = 'blue'
            except Exception as e:
                # self.color = 'red'
                # self.msg = f'ОШИБКА: {type(e)}: {e}'
                color = 'red'
                msg = f'ОШИБКА: {type(e)}: {e}'
                print(msg)
        # else:
        #     self.color = 'red'
        #     self.msg = f'{self.path_to_file}{self.file_name} -- не найден.'

        # return self.data, self.msg, self.color
        return data, msg, color

    def read_comments_from_stock_file_by_openpyxl(self, sheet_name, cell_names_dict):
        comment_dict = dict()
        is_file_path_exists, msg, color = self.is_file_path_exists()
        # print(is_file_path_exists, msg, color)
        # data = None

        if is_file_path_exists:
            path = os.path.abspath(os.path.join(self.file_path, self.file_name))

        # if self.path_to_file is not None and os.path.exists(self.path):
            try:
                # wb = openpyxl.load_workbook(self.path)
                wb = openpyxl.load_workbook(path)
                sheet = wb[sheet_name]
                for cell_col_name, cell_names_list in cell_names_dict.items():
                    cell_names_list = [
    sheet[cell_name].comment.text if sheet[cell_name].comment else None for cell_name in cell_names_list
                    ]
                    comment_dict[sheet[cell_col_name].value] = cell_names_list
                wb.close()
            except Exception as e:
                print(f'ERROR: {type(e)}: {e}')
        # else:
        #     self.msg = f'{self.path} -- not found.'
        #     print(self.msg)
        return comment_dict

    def get_column_names_by_openpyxl(self, sheet_name):
        column_names = []
        is_file_path_exists, msg, color = self.is_file_path_exists()
        if is_file_path_exists:
            path = os.path.abspath(os.path.join(self.file_path, self.file_name))

            try:
                # wb = openpyxl.load_workbook(self.path)
                wb = openpyxl.load_workbook(path)
                sheet = wb[sheet_name]
                column_names = [cell.value for cell in sheet[1]]  # Assumes first row has column names
                wb.close()
            except Exception as e:
                print(f'ERROR: {type(e)}: {e}')

        return column_names

    def get_comments_df_by_openpyxl(self, sheet_name, df, init_index, init_columns):
        print('get_comments_df_by_openpyxl')
        print(df.shape)
        is_file_path_exists, msg, color = self.is_file_path_exists()
        comments_df = pd.DataFrame(index=init_index, columns=init_columns)

        if is_file_path_exists:
            path = os.path.abspath(os.path.join(self.file_path, self.file_name))
            print('try')
            # if self.path_to_file is not None and os.path.exists(self.path):
            try:
                # wb = openpyxl.load_workbook(self.path)
                wb = openpyxl.load_workbook(path)
                sheet = wb[sheet_name]
                # for cell_col_name, cell_names_list in cell_names_dict.items():
                #     cell_names_list = [
                #         sheet[cell_name].comment.text if sheet[cell_name].comment else None for cell_name in
                #         cell_names_list
                #     ]
                #     comment_dict[sheet[cell_col_name].value] = cell_names_list
                for row_idx, row in df.iterrows():
                    print('--------------')
                    print(row_idx, row)
                    for col_idx, value in enumerate(row):
                        print(col_idx, value)
                        if pd.notna(value):  # Check if the cell is not empty
                            # cell = sheet.cell(row=row_idx + 2, column=col_idx + 1)  # Adjust for 1-based index in Excel
                            cell = sheet.cell(row=row_idx + 2,  column=col_idx + 9)
                            # comment = cell.comment
                            if cell.comment:
                                print(cell.comment.text)
                                print(row_idx, col_idx)
                                # 14-8=6  (cols in stock_dev - cols in df = 6)
                                comments_df.iat[row_idx, col_idx+6] = cell.comment.text
                wb.close()
            except Exception as e:
                print(f'ERROR: {type(e)}: {e}')
            print('finish')
            print(comments_df.shape)
            print(comments_df.columns)
        return comments_df

